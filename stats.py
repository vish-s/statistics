
import pandas as pd
import xlrd
import time
from datetime import datetime, timedelta
import datetime as dt
import sys
from openpyxl import Workbook
from openpyxl import load_workbook
from collections import defaultdict
import ConfigLoader


# Set initial Column pivot
col_index = 34
#col_index = 1


def write_to_file(input_file, mean, max, std, median, tile_95, tile_99, sub_five_perc, sub_ten_perc, min, timeouts, total,
                  start_time, end_time, count, column_name):
    """Write stats to the results xslx.
	:param str input_file: Name of the xslx file
	:param float mean: mean of all values in "column_name"
	:param float max: Maximum value in Column "column_name"
	:param float std: Std deviation of values in "column_name"
	:param float median: Median of value in "column_names"
	:param float tile_95: 95 percentile value of values in "column_name"
	:param float tile_99: 99 percentile value of values in "column_name"
	:param float sub_five_perc: % of values below 300 secs in "column_name"
	:param float sub_ten_sec: % of values below 600 secs in "column_name"
	:param float min: Minimum value in "column_name"
	:param int timeouts: number of 'x' or '0' in "column_name"
	:param float total: sum of all values in "column_name"
	:param float start_time: minimum value in "start_time" column
	:param float end_time: Maximum value in "end_time" column
	:param str column_name: Name/title of the Column
	:returns None

    """
    global col_index 
    i = 2

    title = ['Slowest(secs)', 'Mean(secs)', 'Median(secs)', 'Std Deviation(secs)', '95Percentile(secs)',
             '99 Percentile(secs)', 'Sub 5-min(%)', 'Sub 10-min(%)', 'Fastest(secs)', 'Timeouts', 'Total(secs)',
             'Start Time', 'End Time', 'Count']
    data = [max, mean, median, std, tile_95, tile_99, sub_five_perc, sub_ten_perc, min, timeouts, total, start_time,
            end_time, count]

    wb = load_workbook(input_file)
    ws = wb.active
    ws.cell(row=1, column=col_index, value=column_name + "-Summary")

    for values in range(len(data)):
        ws.cell(row=i, column=col_index, value=title[values])
        ws.cell(row=i, column=col_index + 1, value=data[values])
        i += 1

    wb.save(input_file)
    col_index += 3


def stats_for_data_hhmmss(input_file, column_name):
    """Generates stats for data of the form hh:mm:ss in specified "column_name" for a given xslx file
	:param str input_file: Name of the xslx file
	:param str column_name: Name/title of the column
	:returns: None
    """

    df = pd.read_excel(input_file)

    time_list = df[column_name]

    mean = time.strftime('%H:%M:%S', time.gmtime(time_list.astype(str).apply(get_sec).mean()))
    min = time.strftime('%H:%M:%S', time.gmtime(time_list.astype(str).apply(get_sec).min()))
    max = time.strftime('%H:%M:%S', time.gmtime(time_list.astype(str).apply(get_sec).max()))
    std = time.strftime('%H:%M:%S', time.gmtime(time_list.astype(str).apply(get_sec).std()))
    median = time.strftime('%H:%M:%S', time.gmtime(time_list.astype(str).apply(get_sec).median()))
    tile_95 = time.strftime('%H:%M:%S', time.gmtime(time_list.astype(str).apply(get_sec).quantile(0.95)))
    tile_99 = time.strftime('%H:%M:%S', time.gmtime(time_list.astype(str).apply(get_sec).quantile(0.99)))

    print "Mean: {}, Min: {}, Max: {}, Median: {}, Std Deviation: {}, 95 Tile: {} ,99 Tile: {} ".format(mean, min, max,
                                                                                                        median, std,
                                                                                                        tile_95,
                                                                                                        tile_99)
    write_to_file(input_file, mean, min, max, std, median, tile_95, tile_99)


def stats_for_data_secs(input_file, column_name, col_time_dict):
    """
	Generates stats for data of the form float, in a specified "column_name" for a given xslx file
	:param str input_file: Name of the xslx file
	:param str column_name: Name/title of the column
	:returns: None

    """
    # Store data in a xslx file into a Pandas dataframe
    df = pd.read_excel(input_file)

    x = df[column_name] != 0
    config          = ConfigLoader.Config()	
    # Remove rows with 0's and save in new df
    time_list = time_tlist[time_tlist != 0]

    if "Time" not in  column_name:
     	 start_time = df[column_name].min()
     	 end_time = df[column_name].max()
    	 total = df[column_name].max() - df[column_name].min()
     	 time_tlist = df[column_name].dropna().astype(str).astype(float)
     	 time_list = time_tlist[time_tlist != 0]
     	 vsi_count = time_list.count()
     	 start_time = "---"
     	 end_time =   "---"
    else:
    	 start_time = col_time_dict[column_name]['ST']
    	 end_time = col_time_dict[column_name]['ET']
    	 total = col_time_dict[column_name]['TT']
     	 #count = col_time_dict[column_name]['Count']
    	 time_tlist = df[column_name].dropna().astype(str).astype(float)
    	 time_list = time_tlist[time_tlist != 0]
    	 total_count = time_list.count()
    	 start_time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(float(start_time)))
    	 end_time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime(float(end_time)))

    mean_sec = time_list.mean()
    min_sec = time_list.min()
    max_sec = time_list.max()
    std_sec = time_list.std()
    print "std_sec", std_sec
    median_sec = time_list.median()
    tile_95_sec = time_list.quantile(0.95)
    tile_99_sec = time_list.quantile(0.99)
    #total_count = time_tlist.count()

    #tot_count = time_list.count()
    sub_five = df[df < 301][column_name].dropna().astype(str).astype(float).count()
    print sub_five
    sub_ten = df[df < 601][column_name].dropna().astype(str).astype(float).count()
    print sub_ten
    try:
    	 sub_five_perc = float(sub_five) / int(vsi_count) * 100
    except ZeroDivisionError as e:
    	 sub_five_perc = e  
    	 exit()
    try:
    	 sub_ten_perc = float(sub_ten) / int(vsi_count) * 100
    except ZeroDivisionError as e:
    	 sub_ten_perc = e  

    # total = time_list.sum()
    #total = df['End time'].max() - df['Start Time'].min()
    # timeouts = df[column_name].count() - time_list.count()
    timeouts = int(vsi_count) - time_list.count()
     
    mean = time.strftime('%H:%M:%S', time.gmtime(mean_sec))
    min = time.strftime('%H:%M:%S', time.gmtime(min_sec))
    max = time.strftime('%H:%M:%S', time.gmtime(max_sec))
    std = time.strftime('%H:%M:%S', time.gmtime(std_sec))
    median = time.strftime('%H:%M:%S', time.gmtime(median_sec))
    tile_95 = time.strftime('%H:%M:%S', time.gmtime(tile_95_sec))
    tile_99 = time.strftime('%H:%M:%S', time.gmtime(tile_99_sec))

    print "Mean: {}\nMin: {}\nMax: {}\nMedian {}\nStd Deviation: {}\n95 Tile: {}\n99 Tile: {} count: {}\n".format(mean, min, max,
                                                                                                        median, std,
                                                                                                        tile_95,
                                                                                                        tile_99, int(vsi_count))
    print sub_five_perc, sub_ten_perc, total, start_time, end_time

    write_to_file(input_file, mean_sec, max_sec, std_sec, median_sec, tile_95_sec, tile_99_sec, sub_five_perc, sub_ten_perc,
                  min_sec, timeouts, total, start_time, end_time, int(vsi_count),  column_name)


def get_sec(time_str):
    """Convert time in secs to hh:mm:ss format
	:param float time_str: Time in secs
	:returns: Time in the hh:mm:ss format
    """
    h, m, s = time_str.split(':')
    return int(h) * 3600 + int(m) * 60 + int(s)


if __name__ == "__main__":

    input_file = sys.argv[1]
    print input_file
    cols = sys.argv
    for col in cols[2:]:
        column_name = col
        stats_for_data_secs(input_file, column_name)
